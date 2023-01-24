import openpyxl
import requests
from PIL import Image

photo_index = 0


def download_image(link: str):
    global photo_index
    photo_index += 1
    img_data = requests.get(link).content
    with open(f'photos/image_name_{photo_index}.jpg', 'wb') as handler:
        handler.write(img_data)
    handler.close()

    image = Image.open(f'photos/image_name_{photo_index}.jpg')
    sunset_resized = image.resize((250, 250))
    sunset_resized.save(f'photos/image_name_{photo_index}.jpg')
    return f'photos/image_name_{photo_index}.jpg'


def read_file(filename: str):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    ids = []
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, 1):
            ids.append(col[i].value)
    return ids


def write_file(result, filename):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in worksheet['B3:B100000']:
        for cell in row:
            cell.value = None
    diff = 0
    for i in range(1, len(result) + 1):
        try:
            img = openpyxl.drawing.image.Image(download_image(str(result[i - 1][0])))

            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['A'].height = 35
            img.anchor = f'A{i + 2 + diff}'
            worksheet.add_image(img)
            worksheet.cell(row=i + 2 + diff, column=2).value = str(result[i - 1][1])
            worksheet.cell(row=i + 2 + diff, column=3).value = str(result[i - 1][2])
        except:
            pass
        diff_first = -1
        for z in range(len(result[i - 1][3])):
            if z == 0:
                diff_first += 1
            else:
                diff_first += 15
            try:
                img = openpyxl.drawing.image.Image(download_image(str(result[i - 1][3][z])))
                worksheet.column_dimensions['D'].width = 35
                worksheet.column_dimensions['D'].height = 35
                img.anchor = f'D{i + 2 + diff + diff_first}'
                worksheet.add_image(img)
                worksheet.cell(row=i + 2 + diff + diff_first, column=5).value = result[i - 1][4][z]
            except:
                pass

        diff_second = -1
        for z in range(len(result[i - 1][5])):
            if z == 0:
                diff_second += 1
            else:
                diff_second += 15
            try:
                img = openpyxl.drawing.image.Image(download_image(str(result[i - 1][5][z])))
                worksheet.column_dimensions['F'].width = 35
                worksheet.column_dimensions['F'].height = 35
                img.anchor = f'F{i + 2 + diff + diff_second}'
                worksheet.add_image(img)
                worksheet.cell(row=i + 2 + diff + diff_second, column=7).value = result[i - 1][6][z]
            except:
                pass
        diff += max(diff_first, diff_second) + 20  # photo size
    workbook.save("Result.xlsx")
